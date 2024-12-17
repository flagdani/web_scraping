# cndcuritiba.py

import os
import time
import json
import pytesseract
import shutil
import fitz
from PIL import Image
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import undetected_chromedriver as uc
import re
import logging
import pandas as pd

# Configuração do Tesseract
pytesseract.pytesseract.tesseract_cmd = r'path/to/tesseract.exe' # Tesseract para OCR
os.environ['TESSDATA_PREFIX'] = r'path/to/tessdata'  # Tessdata

logging.basicConfig(level=logging.INFO, filename='process.log', filemode='a',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


class ProcessadorPDFCuritiba:
    @staticmethod
    def pixmap_para_pil(pix):
        mode = "RGBA" if pix.alpha else "RGB"
        size = (pix.width, pix.height)
        return Image.frombytes(mode, size, pix.samples)

    def extrair_texto_de_pdf(self, caminho_pdf):
        texto = ""
        with fitz.open(caminho_pdf) as doc:
            for page_number, page in enumerate(doc):
                pix = page.get_pixmap()
                img = self.pixmap_para_pil(pix)
                texto += pytesseract.image_to_string(img, lang='por')
        return texto


class GerenciadorPlanilhaCuritiba:
    def __init__(self, caminho_planilha):
        self.caminho_planilha = caminho_planilha
        self.carregar_planilha()

    def carregar_planilha(self):
        logging.info("Carregando a planilha...")
        self.wb = openpyxl.load_workbook(self.caminho_planilha)
        self.sheet = self.wb.active
        self.indice_coluna_cnpj, self.indice_coluna_cnd_municipal = None, None
        for cell in self.sheet[1]:
            if cell.value == "CNPJ":
                self.indice_coluna_cnpj = cell.column
            elif cell.value == "CND_MUNICIPAL":
                self.indice_coluna_cnd_municipal = cell.column
        if not self.indice_coluna_cnpj or not self.indice_coluna_cnd_municipal:
            logging.error("Colunas 'CNPJ' ou 'CND_MUNICIPAL' não encontradas.")
            exit()

    def salvar_planilha(self):
        logging.info("Salvando a planilha...")
        self.wb.save(self.caminho_planilha)
        logging.info("Planilha salva com sucesso.")

    def atualizar_planilha(self, cnpj, status):
        row_index = self._obter_indice_linha(cnpj)
        if row_index:
            self.sheet.cell(row=row_index, column=self.indice_coluna_cnd_municipal, value=status)
            self.salvar_planilha()

    def _obter_indice_linha(self, cnpj):
        for row in self.sheet.iter_rows(min_row=2):
            if row[self.indice_coluna_cnpj - 1].value == cnpj:
                return row[0].row
        return None


class NavegadorWebCuritiba:
    def __init__(self, lista_cnpjs, dirs, solucionador, processador_pdf, gerenciador_planilha):
        self.lista_cnpjs = lista_cnpjs
        self.dirs = dirs
        self.solucionador = solucionador
        self.processador_pdf = processador_pdf
        self.gerenciador_planilha = gerenciador_planilha

    def configurar_chrome(self):
        chrome_options = uc.ChromeOptions()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--window-size=1920x1080")
        chrome_options.add_argument("--remote-debugging-port=9222")
        chrome_options.add_argument("--log-level=3")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-extensions")

        prefs = {
            "download.default_directory": self.dirs['downloads'],
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        return chrome_options

    def acessar_site(self, driver, cnpj, increase_times=False):
        try:
            driver.get('https://cnd-cidadao.curitiba.pr.gov.br/')
            logging.info(f"Acessando o site para o CNPJ: {cnpj}")
            self._selecionar_certidao(driver, increase_times)
            self._solicitar_certidao(driver, increase_times)
            self._preencher_cnpj(driver, cnpj, increase_times)
            self._resolver_recaptcha(driver, increase_times)
            self._baixar_certidao(driver, cnpj, increase_times)
        except Exception as e:
            if not increase_times:
                logging.warning(
                    f"Primeira tentativa falhou para CNPJ {cnpj}, tentando novamente com tempos aumentados.")
                self.acessar_site(driver, cnpj, increase_times=True)
            else:
                logging.error(f"Ocorreu um erro ao processar o CNPJ {cnpj} com tempos aumentados: {e}")
                self.gerenciador_planilha.atualizar_planilha(cnpj, f"Erro: {e}")

    def _selecionar_certidao(self, driver, increase_times):
        wait_time = 30 if increase_times else 15
        logging.info("Selecionando certidão...")
        dropdown_button = WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Certidão para Pessoa Jurídica/CNPJ')]")))
        dropdown_button.click()

    def _solicitar_certidao(self, driver, increase_times):
        wait_time = 30 if increase_times else 15
        logging.info("Solicitando certidão...")
        solicitar_link = WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@href='/Certidao/SolicitarCnpj']")))
        solicitar_link.click()

    def _preencher_cnpj(self, driver, cnpj, increase_times):
        wait_time = 30 if increase_times else 15
        logging.info(f"Preenchendo CNPJ: {cnpj}")
        cnpj_field = WebDriverWait(driver, wait_time).until(EC.presence_of_element_located((By.ID, 'DocumentoCnpj')))
        time.sleep(5)
        cnpj_field.send_keys(Keys.HOME)
        for char in cnpj:
            cnpj_field.send_keys(char)
            time.sleep(0.2 if increase_times else 0.1)
        time.sleep(5)

    def _resolver_recaptcha(self, driver, increase_times):
        wait_time = 60  # Pode ajustar conforme necessário
        logging.info("Resolvendo reCAPTCHA...")
        site_key = driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')
        url = driver.current_url
        # captcha_response = self.solucionador.resolver_recaptcha(site_key, url)  # Comentado: Implementar solução de CAPTCHA
        captcha_response = "CAPTCHA_RESOLVED"  # Placeholder
        driver.execute_script("document.getElementById('g-recaptcha-response').style.display = 'block';")
        driver.execute_script(
            f"document.getElementById('g-recaptcha-response').value = '{captcha_response}';")
        generate_button = driver.find_element(By.ID, 'btnSolicitar')
        generate_button.click()
        time.sleep(wait_time)

    def _baixar_certidao(self, driver, cnpj, increase_times):
        wait_time = 60  # Pode ajustar conforme necessário
        logging.info(f"Baixando certidão para o CNPJ: {cnpj}")
        generate_new_button = WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.ID, 'btnGerarNovaCertidao')))
        generate_new_button.click()
        time.sleep(wait_time)
        download_button = WebDriverWait(driver, wait_time).until(EC.presence_of_element_located(
            (By.XPATH, "//button[contains(., 'Baixar') and contains(@class, 'btn-primary')]")))
        driver.execute_script("arguments[0].scrollIntoView(true);", download_button)
        time.sleep(5)
        if download_button.is_displayed() and download_button.is_enabled():
            download_button.click()
        else:
            driver.execute_script("arguments[0].click();", download_button)
        time.sleep(wait_time)
        self._processar_pdf_baixado(cnpj)

    def _processar_pdf_baixado(self, cnpj):
        logging.info(f"Processando PDF baixado para o CNPJ: {cnpj}")
        pdf_files = [f for f in os.listdir(self.dirs['downloads']) if f.endswith('.pdf')]
        if not pdf_files:
            logging.error(f"Erro ao baixar PDF para o CNPJ {cnpj}: Nenhum PDF encontrado")
            self.gerenciador_planilha.atualizar_planilha(cnpj, 'Erro: PDF não encontrado')
            return

        pdf_file = max([os.path.join(self.dirs['downloads'], f) for f in pdf_files], key=os.path.getctime)
        novo_caminho_pdf = os.path.join(self.dirs['pdfs'], f"{cnpj}.pdf")
        if os.path.exists(novo_caminho_pdf):
            os.remove(novo_caminho_pdf)
        shutil.move(pdf_file, novo_caminho_pdf)

        texto = self.processador_pdf.extrair_texto_de_pdf(novo_caminho_pdf)
        self._mover_pdf_e_atualizar_planilha(cnpj, texto, novo_caminho_pdf)

    def _mover_pdf_e_atualizar_planilha(self, cnpj, texto, caminho_pdf):
        logging.info(f"Movendo PDF e atualizando planilha para o CNPJ: {cnpj}")
        if "NEGATIVA" in texto.upper() and "POSITIVA COM EFEITO DE NEGATIVA" not in texto.upper():
            novo_dir = self.dirs['negativos']
            status = 'OK, Negativa'
        elif "POSITIVA" in texto.upper():
            if re.search(r"POSITIVA\s*COM\s*EFEITO\s*DE\s*NEGATIVA", texto.upper()):
                novo_dir = self.dirs['positivas_efeito_negativas']
                status = 'OK, Positiva com Efeitos de Negativa'
            else:
                novo_dir = self.dirs['positivos']
                status = 'OK, Positiva'
        else:
            novo_dir = self.dirs['positivos']
            status = 'Texto não reconhecido'

        novo_caminho = os.path.join(novo_dir, f"{cnpj}.pdf")
        shutil.move(caminho_pdf, novo_caminho)
        self.gerenciador_planilha.atualizar_planilha(cnpj, status)
        logging.info(f"PDF movido e planilha atualizada para o CNPJ: {cnpj} com status: {status}")


def carregar_cnpjs(caminho_planilha):
    df = pd.read_excel(caminho_planilha)

    df['CNPJ'] = df['CNPJ'].apply(lambda x: re.sub(r'\D', '', str(x)).zfill(14))
    return df[df['NOME_CIDADE'].str.strip().str.lower() == 'curitiba']['CNPJ'].tolist()


def main():
    caminho_planilha = r'path/to/excel_file.xlsx'  # Caminho da planilha de consulta e salvamento
    dirs = {
        'downloads': os.path.join(os.path.expanduser('~'), 'Downloads'),
        'pdfs': r'path/to/cndcuritiba/pdfs',
        'negativos': r'path/to/cndcuritiba/negativos',
        'positivos': r'path/to/cndcuritiba/positivos',
        'positivas_efeito_negativas': r'path/to/cndcuritiba/positivas_efeito_negativas'
    }
    for dir_path in dirs.values():
        os.makedirs(dir_path, exist_ok=True)

    cnpjs = carregar_cnpjs(caminho_planilha)

    solucionador = None
    processador_pdf = ProcessadorPDFCuritiba()
    gerenciador_planilha = GerenciadorPlanilhaCuritiba(caminho_planilha)
    navegador_web = NavegadorWebCuritiba(cnpjs, dirs, solucionador, processador_pdf, gerenciador_planilha)

    chrome_options = navegador_web.configurar_chrome()
    driver = uc.Chrome(options=chrome_options)

    for cnpj in cnpjs:
        try:
            navegador_web.acessar_site(driver, cnpj)
        except Exception as e:
            logging.error(f"Ocorreu um erro ao processar o CNPJ {cnpj} após tentativa com tempos aumentados: {e}")
            gerenciador_planilha.atualizar_planilha(cnpj, f"Erro: {e}")
        time.sleep(30)

    driver.quit()
    gerenciador_planilha.salvar_planilha()


if __name__ == "__main__":
    main()
