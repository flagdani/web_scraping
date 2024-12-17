# CEF.py

import os
import time
import random
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
import openpyxl
import logging
import base64
import undetected_chromedriver as uc


class ConfiguracaoCEF:
    DOWNLOAD_DIR = r"path/to/downloads"  # Diretório de download dos arquivos
    FINAL_DIR = r"path/to/final_directory"  # Final
    PLANILHA_PATH = r'path/to/excel_file.xlsx'  # Caminho da Planilha de consulta e salvamento

    @staticmethod
    def configurar_logging():
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class PlanilhaCEF:
    def __init__(self, caminho):
        self.caminho = caminho
        self.df = self.carregar_dados()
        self.wb = openpyxl.load_workbook(caminho)
        self.sheet = self.wb.active
        self._obter_indices_colunas()

    def carregar_dados(self):
        logging.info("Carregando dados da planilha")
        df = pd.read_excel(self.caminho, dtype=str)  # Carrega tudo como string
        df['CNPJ'] = df['CNPJ'].apply(self._formatar_cnpj)
        return df

    def _formatar_cnpj(self, cnpj):
        """Remove espaços, hífens e outros caracteres não numéricos, e garante 14 dígitos."""
        cnpj = ''.join(filter(str.isdigit, cnpj))  # Remove tudo que não for número
        return cnpj.zfill(14)  # Garante que tenha 14 dígitos

    def _obter_indices_colunas(self):
        headers = next(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        self.cnpj_index = headers.index("CNPJ")
        self.status_index = headers.index("CEF")

    def obter_cnpjs(self, apenas_erros=False):
        cnpjs = []
        for idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True)):
            cnpj = self._formatar_cnpj(str(row[self.cnpj_index]))  # Formata o CNPJ novamente para garantir
            status = row[self.status_index] if self.status_index < len(row) else ""
            if not apenas_erros or "Erro" in status:
                cnpjs.append((cnpj, idx + 2))
        return cnpjs

    def atualizar_status(self, row_index, status):
        self.sheet.cell(row=row_index, column=self.status_index + 1, value=status)
        self.wb.save(self.caminho)
        logging.info(f"Atualizado status na linha {row_index}: {status}")


class NavegadorCEF:
    def __init__(self):
        self.driver = None

    def iniciar_navegador(self):
        logging.info("Iniciando navegador")
        options = uc.ChromeOptions()
        prefs = {
            "profile.default_content_settings.popups": 0,
            "download.default_directory": ConfiguracaoCEF.DOWNLOAD_DIR,
            "safebrowsing.enabled": "false",
            "profile.default_content_setting_values.automatic_downloads": 1,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.disable_download_protection": True
        }
        options.add_experimental_option("prefs", prefs)
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--no-sandbox')
        self.driver = uc.Chrome(options=options)

    def acessar_site(self, url):
        logging.info(f"Acessando site: {url}")
        self.driver.get(url)
        time.sleep(random.uniform(2, 5))  # Espera aleatória para simular comportamento humano

    def pesquisar_no_google(self, query):
        logging.info(f"Pesquisando no Google: {query}")
        self.driver.get("https://www.google.com")
        self.encontrar_elemento(By.NAME, "q").send_keys(query + "\n")
        time.sleep(random.uniform(2, 5))

    def clicar_primeiro_link(self):
        logging.info("Clicando no primeiro link relevante")
        link = self.encontrar_elemento(By.XPATH, "//a[contains(@href, 'consulta-crf.caixa.gov.br')]")
        link.click()
        time.sleep(random.uniform(2, 5))

    def encontrar_elemento(self, by, value, timeout=20):
        logging.info(f"Procurando elemento: {value}")
        return WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, value)))

    def clicar_elemento(self, by, value, timeout=20):
        logging.info(f"Clicando no elemento: {value}")
        WebDriverWait(self.driver, timeout).until(EC.element_to_be_clickable((by, value))).click()

    def preencher_campo(self, by, value, texto, timeout=20):
        logging.info(f"Preenchendo campo: {value} com {texto}")
        campo = self.encontrar_elemento(by, value, timeout)
        campo.clear()  # Limpa o campo
        campo.send_keys(Keys.HOME)  # Garante que o cursor esteja no início
        time.sleep(1)  # Pequena pausa antes de preencher

        # Digita o texto caractere por caractere com uma pausa entre cada um
        for char in texto:
            campo.send_keys(char)
            time.sleep(random.uniform(0.1, 0.3))  # Pausa aleatória entre as teclas

    def salvar_imagem_captcha(self, xpath, save_path):
        logging.info(f"Salvando imagem do captcha: {xpath}")
        captcha_img = self.encontrar_elemento(By.XPATH, xpath)
        captcha_img.screenshot(save_path)

    def verificar_pendente_sem_pdf(self):
        try:
            feedback = self.driver.find_element(By.XPATH,
                                                "//div[@class='feedback feedback-info']/span[@class='feedback-text']")
            if "As informações disponíveis não são suficientes" in feedback.text:
                return True
            return False
        except NoSuchElementException:
            return False

    def limpar_cache_e_cookies(self):
        logging.info("Limpando cache e cookies")
        self.driver.delete_all_cookies()
        self.driver.execute_cdp_cmd('Network.clearBrowserCache', {})

    def finalizar(self):
        logging.info("Finalizando navegador")
        if self.driver:
            try:
                self.driver.quit()
            except WebDriverException as e:
                logging.error(f"Erro ao finalizar navegador: {e}")
            finally:
                self.driver = None


class SolucionadorCaptchaImg:
    def __init__(self, api_key):
        # self.solver = TwoCaptcha(api_key)  # Comentado: Implementar resolução de CAPTCHA conforme necessário
        pass

    def resolver_captcha(self, image_path):
        logging.info("Resolvendo captcha")
        # Implementar resolução de CAPTCHA aqui, por exemplo, usando uma API de terceiros
        # Exemplo:
        # try:
        #     result = self.solver.normal(image_path, regsense=1)
        #     return result['code']
        # except Exception as e:
        #     logging.error(f"Erro ao resolver CAPTCHA: {e}")
        #     return None
        return "CAPTCHA_RESOLVED"  # Placeholder


class ProcessoCNPJCEF:
    def __init__(self, planilha, solucionador_captcha, navegador):
        self.planilha = planilha
        self.solucionador_captcha = solucionador_captcha
        self.navegador = navegador

    def processar_cnpj(self, cnpj, row_index):
        tentativa = 0
        sucesso = False
        while tentativa < 4 and not sucesso:  # Tentar até 4 vezes
            try:
                tentativa += 1
                logging.info(f"Processando CNPJ {cnpj} - Tentativa {tentativa}")

                # Pesquisar no Google e acessar o site da CEF
                self.navegador.pesquisar_no_google("cnd cef")
                self.navegador.clicar_primeiro_link()

                # Processo no site da CEF
                self.navegador.driver.execute_script("window.scrollTo(0, 500);")
                time.sleep(random.uniform(2, 5))

                self.navegador.preencher_campo(By.ID, "mainForm:txtInscricao1", cnpj)
                captcha_image_path = os.path.join(ConfiguracaoCEF.DOWNLOAD_DIR, "captcha.png")
                self.navegador.salvar_imagem_captcha("//img[@id='captchaImg_N2']", captcha_image_path)
                logging.info("Resolvendo captcha")
                captcha_code = self.solucionador_captcha.resolver_captcha(captcha_image_path)
                if not captcha_code:
                    raise Exception("Falha ao resolver o CAPTCHA")

                # Se o CAPTCHA contiver a palavra "Código", reinicie o navegador
                if "Código" in captcha_code:
                    logging.warning("CAPTCHA contém a palavra 'Código'. Reiniciando navegador...")
                    self.navegador.finalizar()
                    time.sleep(30)
                    self.navegador.iniciar_navegador()
                    self.navegador.limpar_cache_e_cookies()
                    continue  # Reinicia a tentativa

                logging.info(f"Captcha resolvido: {captcha_code}")
                self.navegador.preencher_campo(By.ID, "mainForm:txtCaptcha", captcha_code)
                self.navegador.clicar_elemento(By.ID, "mainForm:btnConsultar")
                time.sleep(10)

                # Verifica se o feedback é "Empregador não cadastrado."
                feedback = self.navegador.driver.find_element(By.XPATH,
                                                              "//div[@class='feedback feedback-info']/span[@class='feedback-text']")
                if "Empregador não cadastrado" in feedback.text:
                    logging.warning(f"CNPJ {cnpj} não está cadastrado.")
                    self.planilha.atualizar_status(row_index, "Empregador não cadastrado")
                    sucesso = True  # Considera como sucesso, pois não é erro
                    return

                # Verifica se o status é "pendente" sem PDF
                if self.navegador.verificar_pendente_sem_pdf():
                    logging.warning(f"Informações insuficientes para o CNPJ {cnpj}.")
                    self.planilha.atualizar_status(row_index, "Ausência")
                    sucesso = True  # Considera como sucesso mesmo que não haja PDF
                    return

                self._baixar_pdf(cnpj, row_index)
                sucesso = True  # Se chegou até aqui, o processamento foi bem-sucedido
            except Exception as e:
                logging.error(f"Tentativa {tentativa} falhou para o CNPJ {cnpj}: {e}")
                if tentativa >= 4:  # Atualiza o status se falhar 4 vezes
                    logging.error(f"Falha ao processar CNPJ {cnpj} após {tentativa} tentativas")
                    self.planilha.atualizar_status(row_index, f"Erro após {tentativa} tentativas")
                else:
                    logging.info(f"Tentando novamente CNPJ {cnpj} (tentativa {tentativa + 1})")
                    time.sleep(random.uniform(5, 10))  # Pausa entre tentativas

    def _baixar_pdf(self, cnpj, row_index):
        try:
            self.navegador.clicar_elemento(By.ID, "mainForm:j_id51")
            self.navegador.clicar_elemento(By.ID, "mainForm:btnVisualizar")
            time.sleep(5)
            self._save_pdf_via_devtools(cnpj, row_index)
        except (NoSuchElementException, TimeoutException, WebDriverException) as e:
            logging.error(f"Erro ao processar resultado para o CNPJ {cnpj}: {e}")
            raise  # Levanta a exceção para tentar novamente

    def _save_pdf_via_devtools(self, cnpj, row_index):
        driver = self.navegador.driver
        logging.info("Salvando PDF via DevTools Protocol")
        try:
            pdf_data = driver.execute_cdp_cmd("Page.printToPDF", {
                "landscape": False,
                "displayHeaderFooter": False,
                "printBackground": True,
                "scale": 1.0,
            })
            pdf_bytes = base64.b64decode(pdf_data['data'])
            pdf_path = os.path.join(ConfiguracaoCEF.FINAL_DIR, f"{cnpj}.pdf")
            with open(pdf_path, 'wb') as f:
                f.write(pdf_bytes)
            logging.info(f"PDF salvo em {pdf_path}")
            self.planilha.atualizar_status(row_index, "Processado com PDF")
        except Exception as e:
            logging.error(f"Erro ao salvar PDF para o CNPJ {cnpj}: {e}")
            self.planilha.atualizar_status(row_index, "Erro ao salvar PDF")


def main():
    ConfiguracaoCEF.configurar_logging()
    planilha = PlanilhaCEF(ConfiguracaoCEF.PLANILHA_PATH)
    solucionador_captcha = SolucionadorCaptchaImg(api_key=None)  # Substitua com a chave de API se implementado
    navegador = NavegadorCEF()

    # Inicia o navegador
    navegador.iniciar_navegador()

    # Cria o processador de CNPJ
    processador = ProcessoCNPJCEF(planilha, solucionador_captcha, navegador)

    # Obtém todos os CNPJs (não filtra por erros)
    cnpjs = planilha.obter_cnpjs()

    # Processa cada CNPJ
    for cnpj, row_index in cnpjs:
        processador.processar_cnpj(cnpj, row_index)

    # Finaliza o navegador
    navegador.finalizar()


if __name__ == "__main__":
    main()
