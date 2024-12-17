# trabalhista.py

import os
import time
import base64
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
# from twocaptcha import TwoCaptcha  # Comentado: Integre uma solução de CAPTCHA conforme necessário
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, InvalidSessionIdException,
    UnexpectedAlertPresentException, WebDriverException, NoSuchWindowException
)
from openpyxl import load_workbook
import shutil

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Caminhos e configuração
input_file = r'path/to/excel_file.xlsx'
negativa_dir = r'path/to/trabalhista_negativa'
positiva_dir = r'path/to/trabalhista_positiva'
downloads_dir = r'path/to/downloads'
api_key = 'your_captcha_api_key'

pytesseract.pytesseract.tesseract_cmd = r'path/to/tesseract.exe' # Tesseract para OCR
os.environ['TESSDATA_PREFIX'] = r'path/to/tessdata'  # Tessdata

def encontrar_arquivo_pdf(diretorio, nome_parcial, timeout=60):
    end_time = time.time() + timeout
    while time.time() < end_time:
        for arquivo in os.listdir(diretorio):
            if arquivo.endswith(".pdf") and nome_parcial in arquivo:
                return os.path.join(diretorio, arquivo)
        time.sleep(5)  # Espera de 5 segundos
    return None

class SolucionadorCaptchaImg:
    def __init__(self, api_key):
        # self.solver = TwoCaptcha(api_key)
        pass

    def resolver_captcha(self, image_path, max_retries=3):
        logging.info("Resolvendo captcha")
        # Implementar resolução de CAPTCHA aqui, por exemplo, usando uma API de terceiros
        # Exemplo:
        # attempts = 0
        # while attempts < max_retries:
        #     try:
        #         result = self.solver.normal(image_path)
        #         return result['code']
        #     except Exception as e:
        #         logging.error(f"Erro ao resolver captcha: {e}")
        #         attempts += 1
        #         time.sleep(5)  # Aguarde antes de tentar novamente
        # raise Exception("Falha ao resolver captcha após várias tentativas.")
        return "CAPTCHA_RESOLVED"

class ProcessadorPDF:
    @staticmethod
    def pixmap_para_pil(pix):
        mode = "RGBA" if pix.alpha else "RGB"
        size = (pix.width, pix.height)
        return Image.frombytes(mode, size, pix.samples)

    @staticmethod
    def extrair_texto_de_pdf(caminho_pdf):
        texto = ""
        with fitz.open(caminho_pdf) as doc:
            for page in doc:
                pix = page.get_pixmap()
                img = ProcessadorPDF.pixmap_para_pil(pix)
                texto += pytesseract.image_to_string(img, lang='por')
        return texto.upper()

class Planilha:
    def __init__(self, caminho):
        self.caminho = caminho
        self.wb = load_workbook(caminho)
        self.sheet = self.wb.active
        self._obter_indices_colunas()

    def _obter_indices_colunas(self):
        headers = next(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        self.cnpj_index = headers.index("CNPJ") + 1
        self.tjus_index = headers.index("TJUS") + 1

    def obter_cnpjs(self, processar_todos=True):
        cnpjs = []
        for row_index, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            cnpj = row[self.cnpj_index - 1]
            tjus_status = row[self.tjus_index - 1]
            # Se for para processar todos ou se o TJUS estiver vazio
            if processar_todos or not tjus_status or not tjus_status.startswith("OK"):
                cnpjs.append((cnpj, row_index))
        return cnpjs

    def atualizar_status(self, row_index, status):
        self.sheet.cell(row=row_index, column=self.tjus_index, value=status)
        self.salvar_planilha()

    def salvar_planilha(self):
        self.wb.save(self.caminho)

def process_cnpj(cnpj, driver, captcha_solver):
    pdf_found = False
    result = "Falhou"  # Resultado padrão caso todas as tentativas falhem

    try:
        # Salvar a guia principal para referência
        main_window = driver.current_window_handle

        # Abrir uma nova guia
        driver.execute_script("window.open('');")
        driver.switch_to.window(driver.window_handles[-1])

        driver.get('https://www.tst.jus.br/certidao1')

        # Aceitar cookies se ainda não foram aceitos
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a.ce-accept.post-accept"))
            ).click()
            driver.refresh()
        except (NoSuchElementException, TimeoutException):
            pass

        # Rolagem para visualizar o iframe e entrar no iframe
        driver.execute_script("window.scrollTo(0, 500);")
        iframe = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "iframe[src*='cndt-certidao']"))
        )
        driver.switch_to.frame(iframe)

        # Clicar no botão "Emitir Certidão" e preencher o CNPJ
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='submit'][value='Emitir Certidão']"))
        ).click()
        cnpj_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[id*='cpfCnpj']"))
        )
        cnpj_input.clear()
        cnpj_input.send_keys(cnpj)

        time.sleep(5)  # Aguardar carregamento do captcha

        try:
            # Verificar se a imagem do captcha está presente
            captcha_image_element = driver.find_element(By.CSS_SELECTOR, "img[alt*='Captcha'], img[id='idImgBase64']")
            captcha_image_src = captcha_image_element.get_attribute("src")
            if 'base64' in captcha_image_src:
                captcha_image_data = captcha_image_src.split(",")[1]
                with open("captcha.png", "wb") as f:
                    f.write(base64.b64decode(captcha_image_data))
                captcha_code = captcha_solver.resolver_captcha("captcha.png")
                captcha_input = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[id='idCampoResposta']"))
                )
                captcha_input.clear()
                captcha_input.send_keys(captcha_code)
            else:
                raise NoSuchElementException("Imagem do captcha não encontrada ou atributo 'src' vazio.")
        except (NoSuchElementException, TimeoutException):
            result = "Falhou: reCAPTCHA detectado"
            driver.delete_all_cookies()  # Limpar cookies e tentar novamente
            return result

        # Clicar no botão "Emitir Certidão" novamente
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[id='gerarCertidaoForm:btnEmitirCertidao']"))
        ).click()

        # Esperar o PDF ser gerado e baixado
        pdf_path = encontrar_arquivo_pdf(downloads_dir, cnpj, timeout=60)
        if pdf_path:
            pdf_found = True
            texto_pdf = ProcessadorPDF.extrair_texto_de_pdf(pdf_path)
            if "NEGATIVA" in texto_pdf:
                destino_pdf = os.path.join(negativa_dir, f"{cnpj}.pdf")
                resultado = "OK, Negativa"
            elif "POSITIVA" in texto_pdf:
                destino_pdf = os.path.join(positiva_dir, f"{cnpj}.pdf")
                resultado = "OK, Positiva"
            else:
                destino_pdf = os.path.join(negativa_dir, f"{cnpj}_NA.pdf")
                resultado = "Falhou: texto desconhecido no PDF"
            shutil.move(pdf_path, destino_pdf)
        else:
            result = "Falhou: PDF não encontrado"

    except UnexpectedAlertPresentException as e:
        try:
            alert = driver.switch_to.alert
            alert_text = alert.text
            logging.warning(f"Alerta inesperado detectado: {alert_text}")
            alert.accept()
            result = "Falhou: alerta inesperado"
        except Exception as inner_exception:
            logging.error(f"Erro ao lidar com o alerta: {inner_exception}")
    except (NoSuchWindowException, WebDriverException) as e:
        logging.error(f"Erro ao processar o CNPJ {cnpj} - Janela ou Sessão perdida: {e}")
        # Reiniciar o driver de maneira controlada
        driver.quit()
        driver = iniciar_driver()  # Reiniciar o driver
        result = "Falhou: janela ou sessão perdida, navegador reiniciado"
    except (NoSuchElementException, TimeoutException, InvalidSessionIdException) as e:
        logging.error(f"Erro ao processar o CNPJ {cnpj}: {e}")
        driver.quit()
        driver = iniciar_driver()  # Reiniciar o driver em caso de erro
        result = "Falhou: erro de sessão"
    finally:
        # Fechar a guia que foi aberta para o processamento
        if len(driver.window_handles) > 1:
            driver.close()
        # Voltar para a guia principal, se existir
        try:
            driver.switch_to.window(main_window)
        except NoSuchWindowException:
            logging.error(f"A janela principal não está mais disponível para o CNPJ {cnpj}")
            driver = iniciar_driver()  # Reiniciar o driver

    return result

def iniciar_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-notifications')  # Desabilitar notificações
    driver = webdriver.Chrome(options=options)
    driver.set_window_size(1920, 1080)
    return driver

if __name__ == "__main__":
    # Perguntar ao usuário se quer processar todos ou apenas com a coluna TJUS vazia
    processar_todos = input("Deseja processar todos os CNPJs? (s/n): ").lower() == 's'

    driver = iniciar_driver()
    captcha_solver = SolucionadorCaptchaImg(api_key=None)
    planilha = Planilha(input_file)

    cnpjs = planilha.obter_cnpjs(processar_todos=processar_todos)

    for cnpj, row_index in cnpjs:
        logging.info(f"Processando CNPJ: {cnpj}")
        status = process_cnpj(cnpj, driver, captcha_solver)
        planilha.atualizar_status(row_index, status)

    # Fechar o navegador no final
    if driver:
        driver.quit()
